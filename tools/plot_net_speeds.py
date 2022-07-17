#!/usr/bin/env python
# Eclipse SUMO, Simulation of Urban MObility; see https://eclipse.org/sumo
# Copyright (C) 2008-2022 German Aerospace Center (DLR) and others.
# This program and the accompanying materials are made available under the
# terms of the Eclipse Public License 2.0 which is available at
# https://www.eclipse.org/legal/epl-2.0/
# This Source Code may also be made available under the following Secondary
# Licenses when the conditions for such availability set forth in the Eclipse
# Public License 2.0 are satisfied: GNU General Public License, version 2
# or later which is available at
# https://www.gnu.org/licenses/old-licenses/gpl-2.0-standalone.html
# SPDX-License-Identifier: EPL-2.0 OR GPL-2.0-or-later

# @file    plot_net_speeds.py
# @author  Daniel Krajzewicz
# @author  Michael Behrisch
# @date    2014-02-19

from __future__ import absolute_import
from __future__ import print_function

import os
import sys
import scipy.io
import numpy
import openpyxl

if 'SUMO_HOME' in os.environ:
    sys.path.append(os.path.join(os.environ['SUMO_HOME'], 'tools'))
else:
    sys.exit("please declare environment variable 'SUMO_HOME'")
import sumolib  # noqa
from sumolib.visualization import helpers  # noqa
import matplotlib.pyplot as plt  # noqa
import matplotlib  # noqa


def main(args=None):
    """The main function; parses options and plots"""
    # ---------- build and read options ----------
    from optparse import OptionParser
    optParser = OptionParser()
    optParser.add_option("-n", "--net", dest="net", metavar="FILE",
                         help="Defines the network to read")
    optParser.add_option("--edge-width", dest="defaultWidth",
                         type="float", default=1, help="Defines the edge width")
    optParser.add_option("--edge-color", dest="defaultColor",
                         default='k', help="Defines the edge color")
    optParser.add_option("--minV", dest="minV",
                         type="float", default=None, help="Define the minimum value boundary")
    optParser.add_option("--maxV", dest="maxV",
                         type="float", default=None, help="Define the maximum value boundary")
    optParser.add_option("-v", "--verbose", dest="verbose", action="store_true",
                         default=False, help="If set, the script says what it's doing")
    # standard plot options
    helpers.addInteractionOptions(optParser)
    helpers.addPlotOptions(optParser)
    # parse
    options, remaining_args = optParser.parse_args(args=args)

    if options.net is None:
        print("Error: a network to load must be given.")
        return 1
    if options.verbose:
        print("Reading network from '%s'" % options.net)
    net = sumolib.net.readNet(options.net)

    
    edgeData = {}
    edgeIDFile = openpyxl.load_workbook("./../test/3PercentClient/edgeDensities-edgeID.xlsx")
    worksheet = edgeIDFile.active
    
    minV = 0
    maxV = 100

    for i in range(0, worksheet.max_row):
        temp = [];
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if isinstance(col[i].value, str):
                temp.append(col[i].value)
            else:
                temp.append(col[i].value*100)
            
        edgeData[temp[0]] = temp[1]
        # minV = min(minV, temp[1])
        # maxV = max(maxV, temp[1])


    # if options.logColors:
#    helpers.logNormalise(colors, maxColorValue)
#  else:
#    helpers.linNormalise(colors, minColorValue, maxColorValue)

    helpers.linNormalise(edgeData, minV, maxV)
    
    for e in edgeData:
        edgeData[e] = helpers.getColor(options, edgeData[e], 1.)
    fig, ax = helpers.openFigure(options)
    ax.set_aspect("equal", None, 'C')
    helpers.plotNet(net, edgeData, {}, options)

    # drawing the legend, at least for the colors
    print("%s -> %s" % (minV, maxV))
    sm = matplotlib.cm.ScalarMappable(
        cmap=matplotlib.cm.get_cmap(options.colormap), norm=matplotlib.colors.Normalize(vmin=minV, vmax=maxV))
    # "fake up the array of the scalar mappable. Urgh..."
    # (pelson, http://stackoverflow.com/questions/8342549/matplotlib-add-colorbar-to-a-sequence-of-line-plots)
    sm._A = []
    plt.colorbar(sm)
    options.nolegend = True
    helpers.closeFigure(fig, ax, options)


if __name__ == "__main__":
    sys.exit(main(sys.argv))
